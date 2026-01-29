"""Initialize ITSG-33 knowledge base with control catalog."""

import json
import os
import sys
from pathlib import Path

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

import chromadb
from chromadb.config import Settings
import openpyxl


def load_controls_from_excel(excel_path: Path) -> list:
    """Load ITSG-33 controls from Excel file."""
    import re

    if not excel_path.exists():
        print(f"Excel file not found: {excel_path}")
        return []

    print(f"Loading controls from: {excel_path}")
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    controls = []

    # Define control families to look for
    control_families = [
        "AC",
        "AT",
        "AU",
        "CA",
        "CM",
        "CP",
        "IA",
        "IR",
        "MA",
        "MP",
        "PE",
        "PL",
        "PS",
        "RA",
        "SA",
        "SC",
        "SI",
    ]

    # Pattern to match control IDs like AC-1, AU-2, etc.
    control_id_pattern = re.compile(r"^([A-Z]{2})-(\d+)")

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip().lower())
            else:
                headers.append("")

        print(f"Processing sheet: {sheet_name} with {len(headers)} columns")

        # Process data rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue

            first_cell = str(row[0]).strip() if row[0] else ""
            second_cell = str(row[1]).strip() if len(row) > 1 and row[1] else ""

            # Try to find control ID - could be in first or second column
            control_id = None
            control_name = ""
            family = ""

            # Check if first cell is a control ID
            match = control_id_pattern.match(first_cell.upper())
            if match:
                control_id = first_cell.upper()
                family = match.group(1)
                control_name = second_cell
            else:
                # Check if second cell contains a control ID (e.g., "AC-1 Access Control Policy")
                match = control_id_pattern.match(second_cell.upper())
                if match:
                    control_id = match.group(0)
                    family = match.group(1)
                    control_name = second_cell
                else:
                    # Try to extract control ID from anywhere in first two cells
                    for cell_text in [first_cell, second_cell]:
                        match = control_id_pattern.search(cell_text.upper())
                        if match:
                            control_id = match.group(0)
                            family = match.group(1)
                            control_name = cell_text
                            break

            if not control_id or not family:
                continue

            # Build control record
            control = {
                "id": control_id,
                "family": family,
                "name": control_name,
                "description": "",
                "profile": 1,
                "questions": [],
            }

            # Look for description and questions in other columns
            for i, header in enumerate(headers):
                if i >= len(row):
                    break
                value = row[i]
                if not value:
                    continue

                value_str = str(value).strip()

                if "description" in header or "requirement" in header:
                    control["description"] = value_str
                elif "profile" in header or "level" in header:
                    try:
                        control["profile"] = int(value_str)
                    except (ValueError, TypeError):
                        pass
                elif "question" in header:
                    if value_str:
                        control["questions"].append(value_str)

            # If no description found, use the name
            if not control["description"] and control["name"]:
                control["description"] = control["name"]

            controls.append(control)

    print(f"Loaded {len(controls)} controls from Excel")
    return controls


def load_controls_from_json(json_path: Path) -> list:
    """Load controls from JSON file."""
    if not json_path.exists():
        return []

    with open(json_path) as f:
        return json.load(f)


def create_default_controls() -> list:
    """Create a default set of ITSG-33 controls."""
    controls = []

    # Access Control (AC)
    ac_controls = [
        ("AC-1", "Access Control Policy and Procedures", 1),
        ("AC-2", "Account Management", 1),
        ("AC-3", "Access Enforcement", 1),
        ("AC-4", "Information Flow Enforcement", 2),
        ("AC-5", "Separation of Duties", 2),
        ("AC-6", "Least Privilege", 2),
        ("AC-7", "Unsuccessful Logon Attempts", 1),
        ("AC-8", "System Use Notification", 1),
        ("AC-11", "Session Lock", 2),
        ("AC-12", "Session Termination", 2),
        ("AC-14", "Permitted Actions without Identification or Authentication", 1),
        ("AC-17", "Remote Access", 1),
        ("AC-18", "Wireless Access", 1),
        ("AC-19", "Access Control for Mobile Devices", 1),
        ("AC-20", "Use of External Information Systems", 1),
        ("AC-21", "Information Sharing", 2),
    ]

    for ctrl_id, name, profile in ac_controls:
        controls.append(
            {
                "id": ctrl_id,
                "family": "AC",
                "name": name,
                "description": f"{name} - Establishes requirements for {name.lower()}.",
                "profile": profile,
                "questions": [f"Is {name.lower()} implemented?"],
            }
        )

    # Audit and Accountability (AU)
    au_controls = [
        ("AU-1", "Audit and Accountability Policy and Procedures", 1),
        ("AU-2", "Audit Events", 1),
        ("AU-3", "Content of Audit Records", 1),
        ("AU-4", "Audit Storage Capacity", 1),
        ("AU-5", "Response to Audit Processing Failures", 1),
        ("AU-6", "Audit Review, Analysis, and Reporting", 1),
        ("AU-7", "Audit Reduction and Report Generation", 2),
        ("AU-8", "Time Stamps", 1),
        ("AU-9", "Protection of Audit Information", 1),
        ("AU-11", "Audit Record Retention", 1),
        ("AU-12", "Audit Generation", 1),
    ]

    for ctrl_id, name, profile in au_controls:
        controls.append(
            {
                "id": ctrl_id,
                "family": "AU",
                "name": name,
                "description": f"{name} - {name} requirements.",
                "profile": profile,
                "questions": [f"Is {name.lower()} implemented?"],
            }
        )

    # Add more families...
    families = {
        "AT": [
            ("AT-1", "Security Awareness and Training Policy", 1),
            ("AT-2", "Security Awareness Training", 1),
            ("AT-3", "Role-Based Security Training", 1),
        ],
        "CA": [
            ("CA-1", "Security Assessment and Authorization Policy", 1),
            ("CA-2", "Security Assessments", 1),
            ("CA-3", "System Interconnections", 1),
            ("CA-5", "Plan of Action and Milestones", 1),
            ("CA-6", "Security Authorization", 1),
            ("CA-7", "Continuous Monitoring", 1),
        ],
        "CM": [
            ("CM-1", "Configuration Management Policy", 1),
            ("CM-2", "Baseline Configuration", 1),
            ("CM-3", "Configuration Change Control", 2),
            ("CM-4", "Security Impact Analysis", 1),
            ("CM-5", "Access Restrictions for Change", 1),
            ("CM-6", "Configuration Settings", 1),
            ("CM-7", "Least Functionality", 1),
            ("CM-8", "Information System Component Inventory", 1),
        ],
        "CP": [
            ("CP-1", "Contingency Planning Policy", 1),
            ("CP-2", "Contingency Plan", 1),
            ("CP-3", "Contingency Training", 1),
            ("CP-4", "Contingency Plan Testing", 1),
            ("CP-9", "Information System Backup", 1),
            ("CP-10", "Information System Recovery and Reconstitution", 1),
        ],
        "IA": [
            ("IA-1", "Identification and Authentication Policy", 1),
            ("IA-2", "Identification and Authentication (Organizational Users)", 1),
            ("IA-4", "Identifier Management", 1),
            ("IA-5", "Authenticator Management", 1),
            ("IA-6", "Authenticator Feedback", 1),
            ("IA-7", "Cryptographic Module Authentication", 1),
            ("IA-8", "Identification and Authentication (Non-Organizational Users)", 1),
        ],
        "IR": [
            ("IR-1", "Incident Response Policy", 1),
            ("IR-2", "Incident Response Training", 1),
            ("IR-4", "Incident Handling", 1),
            ("IR-5", "Incident Monitoring", 1),
            ("IR-6", "Incident Reporting", 1),
            ("IR-7", "Incident Response Assistance", 1),
            ("IR-8", "Incident Response Plan", 1),
        ],
        "MA": [
            ("MA-1", "System Maintenance Policy", 1),
            ("MA-2", "Controlled Maintenance", 1),
            ("MA-3", "Maintenance Tools", 1),
            ("MA-4", "Nonlocal Maintenance", 1),
            ("MA-5", "Maintenance Personnel", 1),
        ],
        "MP": [
            ("MP-1", "Media Protection Policy", 1),
            ("MP-2", "Media Access", 1),
            ("MP-6", "Media Sanitization", 1),
            ("MP-7", "Media Use", 1),
        ],
        "PE": [
            ("PE-1", "Physical and Environmental Protection Policy", 1),
            ("PE-2", "Physical Access Authorizations", 1),
            ("PE-3", "Physical Access Control", 1),
            ("PE-6", "Monitoring Physical Access", 1),
            ("PE-8", "Visitor Access Records", 1),
        ],
        "PL": [
            ("PL-1", "Security Planning Policy", 1),
            ("PL-2", "System Security Plan", 1),
            ("PL-4", "Rules of Behavior", 1),
        ],
        "PS": [
            ("PS-1", "Personnel Security Policy", 1),
            ("PS-2", "Position Risk Designation", 1),
            ("PS-3", "Personnel Screening", 1),
            ("PS-4", "Personnel Termination", 1),
            ("PS-5", "Personnel Transfer", 1),
            ("PS-6", "Access Agreements", 1),
            ("PS-7", "Third-Party Personnel Security", 1),
            ("PS-8", "Personnel Sanctions", 1),
        ],
        "RA": [
            ("RA-1", "Risk Assessment Policy", 1),
            ("RA-2", "Security Categorization", 1),
            ("RA-3", "Risk Assessment", 1),
            ("RA-5", "Vulnerability Scanning", 1),
        ],
        "SA": [
            ("SA-1", "System and Services Acquisition Policy", 1),
            ("SA-2", "Allocation of Resources", 1),
            ("SA-3", "System Development Life Cycle", 1),
            ("SA-4", "Acquisition Process", 1),
            ("SA-5", "Information System Documentation", 1),
            ("SA-8", "Security Engineering Principles", 1),
            ("SA-9", "External Information System Services", 1),
        ],
        "SC": [
            ("SC-1", "System and Communications Protection Policy", 1),
            ("SC-5", "Denial of Service Protection", 1),
            ("SC-7", "Boundary Protection", 1),
            ("SC-12", "Cryptographic Key Establishment and Management", 1),
            ("SC-13", "Cryptographic Protection", 1),
            ("SC-15", "Collaborative Computing Devices", 1),
            ("SC-20", "Secure Name/Address Resolution Service", 1),
            ("SC-21", "Secure Name/Address Resolution Service (Recursive)", 1),
            ("SC-22", "Architecture for Name/Address Resolution Service", 1),
        ],
        "SI": [
            ("SI-1", "System and Information Integrity Policy", 1),
            ("SI-2", "Flaw Remediation", 1),
            ("SI-3", "Malicious Code Protection", 1),
            ("SI-4", "Information System Monitoring", 1),
            ("SI-5", "Security Alerts and Advisories", 1),
            ("SI-12", "Information Handling and Retention", 1),
        ],
    }

    for family, family_controls in families.items():
        for ctrl_id, name, profile in family_controls:
            controls.append(
                {
                    "id": ctrl_id,
                    "family": family,
                    "name": name,
                    "description": f"{name} - Establishes requirements for {name.lower()}.",
                    "profile": profile,
                    "questions": [f"Is {name.lower()} implemented?"],
                }
            )

    return controls


def initialize_vector_db(controls: list, persist_dir: str = "./chroma_db"):
    """Initialize ChromaDB with ITSG-33 controls."""
    print(f"Initializing ChromaDB at: {persist_dir}")

    client = chromadb.Client(Settings(persist_directory=persist_dir, anonymized_telemetry=False))

    # Delete existing collection if exists
    try:
        client.delete_collection("itsg33_controls")
        print("Deleted existing collection")
    except Exception:
        pass

    collection = client.create_collection(
        name="itsg33_controls", metadata={"description": "ITSG-33 security controls catalog"}
    )

    print(f"Adding {len(controls)} controls to vector database...")

    # Add controls in batches
    batch_size = 100
    for i in range(0, len(controls), batch_size):
        batch = controls[i : i + batch_size]

        ids = []
        documents = []
        metadatas = []

        for c in batch:
            # Index English version
            ids.append(f"{c['id']}_en")
            documents.append(
                f"{c.get('name_en', c.get('name'))}. {c.get('description_en', c.get('description'))}"
            )
            metadatas.append(
                {
                    "id": c["id"],
                    "family": c["family"],
                    "profile": c["profile"],
                    "lang": "en",
                    "name": c.get("name_en", c.get("name")),
                    "questions": json.dumps(c.get("questions_en", c.get("questions", []))),
                }
            )

            # Index French version if available
            if "name_fr" in c:
                ids.append(f"{c['id']}_fr")
                documents.append(f"{c['name_fr']}. {c['description_fr']}")
                metadatas.append(
                    {
                        "id": c["id"],
                        "family": c["family"],
                        "profile": c["profile"],
                        "lang": "fr",
                        "name": c["name_fr"],
                        "questions": json.dumps(c.get("questions_fr", [])),
                    }
                )

        collection.add(ids=ids, documents=documents, metadatas=metadatas)

        print(f"Added batch {i // batch_size + 1}: {len(batch)} controls")

    print(f"Successfully loaded {len(controls)} controls into vector database")
    return collection


def save_controls_to_json(controls: list, output_path: Path):
    """Save controls to JSON file."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w") as f:
        json.dump(controls, f, indent=2)
    print(f"Saved controls to: {output_path}")


def main():
    """Main setup function."""
    project_root = Path(__file__).parent.parent
    data_dir = project_root / "data"
    data_dir.mkdir(exist_ok=True)

    # Try to load from Excel first
    excel_path = project_root.parent / "ITSG-33 Security Control Questions.xlsx"
    controls = []

    if excel_path.exists():
        controls = load_controls_from_excel(excel_path)

    # If no controls from Excel, try bilingual JSON first, then standard JSON
    if not controls:
        bilingual_path = data_dir / "itsg33_controls_bilingual.json"
        if bilingual_path.exists():
            controls = load_controls_from_json(bilingual_path)
            print(f"Loaded {len(controls)} bilingual controls from JSON")
        else:
            json_path = data_dir / "itsg33_controls.json"
            controls = load_controls_from_json(json_path)
            if controls:
                print(f"Loaded {len(controls)} controls from standard JSON")

    # If still no controls, create defaults
    if not controls:
        print("Creating default control catalog...")
        controls = create_default_controls()

    # Save to JSON for future use
    save_controls_to_json(controls, data_dir / "itsg33_controls.json")

    # Initialize vector database
    chroma_dir = os.getenv("CHROMA_PERSIST_DIR", str(project_root / "chroma_db"))
    initialize_vector_db(controls, chroma_dir)

    # Create profiles.json
    profiles = [
        {
            "number": 1,
            "name": "Profile 1 - Low",
            "description": "For systems with low sensitivity data and low impact",
            "control_count": len([c for c in controls if c["profile"] <= 1]),
        },
        {
            "number": 2,
            "name": "Profile 2 - Moderate",
            "description": "For systems with moderate sensitivity data",
            "control_count": len([c for c in controls if c["profile"] <= 2]),
        },
        {
            "number": 3,
            "name": "Profile 3 - High",
            "description": "For systems with high sensitivity data or critical operations",
            "control_count": len(controls),
        },
    ]

    with open(data_dir / "profiles.json", "w") as f:
        json.dump(profiles, f, indent=2)
    print(f"Saved profiles to: {data_dir / 'profiles.json'}")

    print("\nSetup complete!")
    print(f"  - Controls: {len(controls)}")
    print(f"  - Data directory: {data_dir}")
    print(f"  - ChromaDB directory: {chroma_dir}")


if __name__ == "__main__":
    main()

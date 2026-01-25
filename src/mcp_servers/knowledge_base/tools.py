"""Additional tools for Knowledge Base MCP Server."""

import json
from pathlib import Path
from typing import List, Dict, Any, Optional
import openpyxl


def load_controls_from_excel(excel_path: Path) -> List[Dict[str, Any]]:
    """
    Load ITSG-33 controls from Excel file.

    Args:
        excel_path: Path to the Excel file

    Returns:
        List of control dictionaries
    """
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    controls = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Skip if not a control family sheet
        if sheet_name not in [
            "AC", "AT", "AU", "CA", "CM", "CP", "IA", "IR",
            "MA", "MP", "PE", "PL", "PS", "RA", "SA", "SC", "SI"
        ]:
            continue

        # Process rows (assuming header in row 1)
        headers = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue

            control = {
                "id": str(row[0]) if row[0] else "",
                "family": sheet_name,
                "name": str(row[1]) if len(row) > 1 and row[1] else "",
                "description": str(row[2]) if len(row) > 2 and row[2] else "",
                "profile": int(row[3]) if len(row) > 3 and row[3] else 1,
                "questions": [],
            }

            # Extract questions if present
            for i, header in enumerate(headers):
                if header and "question" in str(header).lower() and len(row) > i:
                    if row[i]:
                        control["questions"].append(str(row[i]))

            controls.append(control)

    return controls


def convert_excel_to_json(excel_path: Path, output_path: Path) -> None:
    """
    Convert Excel control catalog to JSON format.

    Args:
        excel_path: Path to input Excel file
        output_path: Path to output JSON file
    """
    controls = load_controls_from_excel(excel_path)

    with open(output_path, "w") as f:
        json.dump(controls, f, indent=2)


def get_family_name(family_code: str) -> str:
    """Get full name for a control family code."""
    family_names = {
        "AC": "Access Control",
        "AT": "Awareness and Training",
        "AU": "Audit and Accountability",
        "CA": "Assessment, Authorization, and Monitoring",
        "CM": "Configuration Management",
        "CP": "Contingency Planning",
        "IA": "Identification and Authentication",
        "IR": "Incident Response",
        "MA": "Maintenance",
        "MP": "Media Protection",
        "PE": "Physical and Environmental Protection",
        "PL": "Planning",
        "PS": "Personnel Security",
        "RA": "Risk Assessment",
        "SA": "System and Services Acquisition",
        "SC": "System and Communications Protection",
        "SI": "System and Information Integrity",
    }
    return family_names.get(family_code.upper(), family_code)
